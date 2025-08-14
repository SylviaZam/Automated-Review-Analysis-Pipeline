#!/usr/bin/env python3
"""
Analyze customer survey answers and export a wide-format Excel report
with polished per-product pie charts (one pie per question).

Output:
- One worksheet per Product (wide columns: Qx_Answer / Qx_Sentiment / Qx_Category),
  with wrapped text and auto-sized columns.
- One worksheet "Charts - <Product>" containing a grid of pie charts, one per question,
  showing percentages of Positive / Neutral / Negative / Mixed.
- A "Summary" worksheet aggregating counts per Product × Question × Sentiment.

Modes:
- Demo Mode (no OPENAI_API_KEY): offline analysis via VADER + keyword rules.
- OpenAI Mode (OPENAI_API_KEY set): calls OpenAI for sentiment/category.
"""

import argparse
import json
import os
import re
import sys
import time
from typing import Dict, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv

# Optional: language detection (info log only)
try:
    from langdetect import detect  # type: ignore
except Exception:
    detect = None

# Optional: OpenAI SDK
try:
    from openai import OpenAI  # type: ignore
except Exception:
    OpenAI = None

# Offline sentiment for Demo Mode
try:
    from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer  # type: ignore
    _VADER_ANALYZER = SentimentIntensityAnalyzer()
except Exception:
    _VADER_ANALYZER = None

# openpyxl for formatting & charts
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.workbook.workbook import Workbook

# Treat these as "no feedback"
FILLER_VALUES = {
    "", "n/a", "na", "no", "none", "null", "nan", "sin comentarios", "ninguno", "-", " "
}

# Simple keyword categories for Demo Mode (EN + ES)
DEMO_KEYWORDS = [
    ("Price",    ["price", "expensive", "too expensive", "cheap", "cost", "pricing", "value", "caro", "barato", "precio"]),
    ("Shipping", ["ship", "shipping", "delivery", "arrive", "delay", "delayed", "late", "envío", "envio", "tarde", "demor", "entrega"]),
    ("Quality",  ["quality", "material", "durable", "break", "defect", "defecto", "calidad"]),
    ("Fit",      ["fit", "size", "sizing", "tight", "loose", "talla", "ajuste", "grande", "chico"]),
    ("Design",   ["design", "style", "color", "look", "diseño", "estilo", "colores"]),
    ("Support",  ["support", "help", "service", "refund", "return", "soporte", "atención", "atencion", "reembolso", "devolución", "devolucion"]),
]


# ---------------- Basic helpers ----------------

def clean_text(s: str) -> str:
    """Trim, remove astral-plane emoji, collapse whitespace."""
    if not isinstance(s, str):
        return ""
    s = s.strip()
    s = re.sub(r"[\U00010000-\U0010ffff]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def is_filler(s: str) -> bool:
    t = (s or "").strip().lower()
    return t in FILLER_VALUES


def get_question_columns(df: pd.DataFrame) -> List[str]:
    """Assume first three columns are Email, Name, Products; rest are questions."""
    return list(df.columns[3:]) if df.shape[1] > 3 else []


def detect_language(sample_answers: List[str]) -> Optional[str]:
    for a in sample_answers:
        a = clean_text(a)
        if a and detect:
            try:
                return detect(a)
            except Exception:
                pass
    return None


def normalize_sentiment(s: str) -> str:
    mapping = {"positive": "Positive", "neutral": "Neutral", "negative": "Negative", "mixed": "Mixed"}
    t = (s or "").strip().lower()
    return mapping.get(t, "Neutral")


# ---------------- Demo-mode analyzer ----------------

def _demo_category(answer_lower: str) -> str:
    for cat, kws in DEMO_KEYWORDS:
        if any(k in answer_lower for k in kws):
            return cat
    return "General"


def _demo_sentiment(answer_text: str, answer_lower: str) -> str:
    if _VADER_ANALYZER is not None:
        try:
            score = _VADER_ANALYZER.polarity_scores(answer_text)["compound"]
            if score >= 0.35:
                return "Positive"
            if score <= -0.35:
                return "Negative"
            if any(w in answer_lower for w in ["but", "aunque", "pero"]) and abs(score) < 0.35:
                return "Mixed"
            return "Neutral"
        except Exception:
            pass
    # Tiny fallback
    pos = ["love", "loved", "great", "good", "excellent", "amazing", "encanta", "bueno", "genial", "excelente"]
    neg = ["bad", "poor", "terrible", "awful", "hate", "malo", "expensive", "too expensive", "caro", "tarde", "defecto", "delay", "delayed", "late"]
    p = sum(w in answer_lower for w in pos)
    n = sum(w in answer_lower for w in neg)
    if p and n:
        return "Mixed"
    if p:
        return "Positive"
    if n:
        return "Negative"
    return "Neutral"


def demo_analyze_answer(answer: str) -> Tuple[str, str]:
    text = (answer or "").strip()
    low = text.lower()
    return _demo_sentiment(text, low), _demo_category(low)


# ---------------- OpenAI path ----------------

def call_openai_analyze(industry: str, question: str, answer: str, client: "OpenAI", model: str = "gpt-4o-mini") -> Tuple[str, str]:
    """Ask OpenAI for JSON {sentiment, category}. Returns (sentiment, category)."""
    sys_prompt = "You are an assistant that analyzes customer feedback."
    user_prompt = (
        "Respond ONLY as JSON with keys 'sentiment' and 'category'.\n"
        f"Industry: {industry}\nQuestion: {question}\nAnswer: {answer}\n"
        "Sentiment must be one of: Positive, Neutral, Negative, Mixed. Category should be 1 to 3 words."
    )
    delay = 1.0
    for attempt in range(4):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[{"role": "system", "content": sys_prompt},
                          {"role": "user", "content": user_prompt}],
                temperature=0.2,
            )
            content = resp.choices[0].message.content or "{}"
            m = re.search(r"\{.*\}", content, re.S)
            payload = json.loads(m.group(0) if m else content)
            sentiment = normalize_sentiment(str(payload.get("sentiment", "Neutral")))
            category = (payload.get("category") or "No Feedback").strip()
            if sentiment not in {"Positive", "Neutral", "Negative", "Mixed"}:
                sentiment = "Neutral"
            if not category:
                category = "No Feedback"
            return sentiment, category
        except Exception as e:
            if attempt == 3:
                print(f"[warn] OpenAI failed: {e}. Defaulting to Neutral/No Feedback.", file=sys.stderr)
                return "Neutral", "No Feedback"
            time.sleep(delay)
            delay *= 2
    return "Neutral", "No Feedback"


# ---------------- Data shaping ----------------

def analyze_dataframe_wide(df: pd.DataFrame, industry: str, client: Optional["OpenAI"]) -> pd.DataFrame:
    """Build wide-format table with Qx_Answer, Qx_Sentiment, Qx_Category."""
    results: List[Dict[str, str]] = []
    qcols = get_question_columns(df)

    # Optional language info
    if qcols:
        samples = []
        for q in qcols:
            s = df[q].dropna()
            if not s.empty:
                samples.append(str(s.iloc[0]))
        lang = detect_language(samples)
        if lang:
            print(f"[info] Detected language: {lang}")

    for idx, row in df.iterrows():
        products_raw = str(row.get(df.columns[2], "")).strip()
        products = [p.strip() for p in products_raw.split(",") if p.strip()] or ["Unspecified"]

        # Analyze each question for this response
        q_triplets: Dict[str, Tuple[str, str, str]] = {}
        for q in qcols:
            ans = clean_text(str(row.get(q, "")))
            if is_filler(ans):
                sent, cat = "Neutral", "No Feedback"
            else:
                if client is None:
                    sent, cat = demo_analyze_answer(ans)
                else:
                    sent, cat = call_openai_analyze(industry, q, ans, client)
            q_triplets[q] = (ans, sent, cat)

        # Emit one row per product
        for prod in products:
            out: Dict[str, str] = {"ResponseID": str(idx + 1), "Product": prod[:100]}
            for q in qcols:
                base = re.sub(r"\s+", "_", str(q).strip())
                ans, sent, cat = q_triplets[q]
                out[f"{base}_Answer"] = ans
                out[f"{base}_Sentiment"] = sent
                out[f"{base}_Category"] = cat
            results.append(out)

    if not results:
        return pd.DataFrame(columns=["Product"])

    # Column order
    cols = ["ResponseID", "Product"]
    for q in qcols:
        base = re.sub(r"\s+", "_", str(q).strip())
        cols += [f"{base}_Answer", f"{base}_Sentiment", f"{base}_Category"]

    wide = pd.DataFrame(results)
    existing = [c for c in cols if c in wide.columns]
    remainder = [c for c in wide.columns if c not in existing]
    return wide[existing + remainder]


def build_summary_from_wide(wide: pd.DataFrame) -> pd.DataFrame:
    """Aggregate sentiment counts per Product × Question."""
    if wide.empty:
        return pd.DataFrame()

    # Identify base question names from *_Sentiment columns
    q_names = []
    for c in wide.columns:
        if c.endswith("_Sentiment"):
            q_names.append(c[:-len("_Sentiment")])

    rows = []
    for _, r in wide.iterrows():
        product = r.get("Product", "Unspecified")
        for base in q_names:
            sentiment = str(r.get(f"{base}_Sentiment", "")).strip() or "Neutral"
            rows.append({"Product": product, "Question": base, "Sentiment": sentiment})

    long_df = pd.DataFrame(rows)
    if long_df.empty:
        return pd.DataFrame()

    grouped = long_df.groupby(["Product", "Question", "Sentiment"]).size().reset_index(name="Count")
    pivot = grouped.pivot_table(
        index=["Product", "Question"],
        columns="Sentiment",
        values="Count",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    for s in ["Positive", "Neutral", "Negative", "Mixed"]:
        if s not in pivot.columns:
            pivot[s] = 0

    ordered = ["Product", "Question", "Positive", "Neutral", "Negative", "Mixed"]
    existing = [c for c in ordered if c in pivot.columns]
    remainder = [c for c in pivot.columns if c not in existing]
    return pivot[existing + remainder]


def sanitize_sheet_name(name: str) -> str:
    """Excel sheet name constraints: <=31 chars, no []:*?/\\."""
    s = re.sub(r"[:\\/?*\[\]]", " ", name)
    return s[:31] or "Sheet"


# ---------------- Formatting helpers ----------------

def _autofit_and_wrap(ws, wrap_answer_columns_only: bool = True):
    """Auto-size columns and wrap text for *_Answer columns."""
    header = [cell.value if cell.value is not None else "" for cell in ws[1]]
    answer_col_idx = set(i for i, h in enumerate(header, start=1) if isinstance(h, str) and h.endswith("_Answer"))

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            v = row[0].value
            l = len(str(v)) if v is not None else 0
            max_len = max(max_len, l)

            if (not wrap_answer_columns_only) or (col_idx in answer_col_idx):
                row[0].alignment = Alignment(wrap_text=True, vertical="top")

        width = min(max(12, int(max_len * 0.9)), 60)
        ws.column_dimensions[col_letter].width = width


def _write_product_pie_charts(wb: Workbook, product: str, prod_summary_df: pd.DataFrame, charts_per_row: int = 2):
    """
    Create a new sheet "Charts - <Product>" with a grid of pie charts:
    one pie per question, showing sentiment proportions.
    """
    title = f"Charts - {product}"
    sheet_name = sanitize_sheet_name(title)
    base = sheet_name
    i = 1
    while sheet_name in wb.sheetnames:
        sheet_name = sanitize_sheet_name(f"{base} ({i})")
        i += 1
    ws = wb.create_sheet(sheet_name)

    # Header row with sentiment labels (for consistent categories)
    headers = ["Question", "Positive", "Neutral", "Negative", "Mixed"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)

    # Write the compact table that drives the charts
    for r_idx, (_, row) in enumerate(prod_summary_df.iterrows(), start=2):
        ws.cell(row=r_idx, column=1, value=row["Question"])
        ws.cell(row=r_idx, column=2, value=int(row.get("Positive", 0)))
        ws.cell(row=r_idx, column=3, value=int(row.get("Neutral", 0)))
        ws.cell(row=r_idx, column=4, value=int(row.get("Negative", 0)))
        ws.cell(row=r_idx, column=5, value=int(row.get("Mixed", 0)))

    # Create one pie per row (question)
    num_rows = len(prod_summary_df)
    chart_w, chart_h = 17, 12  # inches*10 roughly; good proportions in Excel
    start_col_letters = ["H", "O", "V", "AC"]  # supports up to 4 charts per row if needed

    for idx in range(num_rows):
        data_row = 2 + idx
        # Data: the row for this question (counts)
        data = Reference(ws, min_col=2, max_col=5, min_row=data_row, max_row=data_row)
        # Categories: the header with sentiment names
        cats = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=1)
        total = sum(int(ws.cell(row=data_row, column=c).value or 0) for c in range(2, 6))
        q_label = str(ws.cell(row=data_row, column=1).value)

        pie = PieChart()
        pie.title = f"{q_label} – Sentiment Mix (n={total})"
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True

        # Place in a grid
        row_block = idx // charts_per_row
        col_block = idx % charts_per_row
        anchor_col = start_col_letters[min(col_block, len(start_col_letters)-1)]
        anchor_row = 2 + row_block * 18  # vertical spacing between rows of charts
        ws.add_chart(pie, f"{anchor_col}{anchor_row}")
        pie.width = chart_w
        pie.height = chart_h

    _autofit_and_wrap(ws, wrap_answer_columns_only=False)


# ---------------- Excel writer ----------------

def write_excel_wide(wide: pd.DataFrame, out_path: str) -> None:
    summary_all = build_summary_from_wide(wide)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        wb = writer.book

        # Per-product data sheets
        if not wide.empty:
            for prod, sub in wide.groupby("Product"):
                sheet = sanitize_sheet_name(str(prod))
                sub_sorted = sub.sort_values("ResponseID")
                sub_sorted.to_excel(writer, index=False, sheet_name=sheet)

        # Summary sheet
        if summary_all is not None and not summary_all.empty:
            summary_all.to_excel(writer, index=False, sheet_name="Summary")

        # Format product sheets
        for prod in wide["Product"].unique() if not wide.empty else []:
            sheet = sanitize_sheet_name(str(prod))
            if sheet in writer.sheets:
                ws = writer.sheets[sheet]
                _autofit_and_wrap(ws, wrap_answer_columns_only=True)

        # Add pie charts per product
        if summary_all is not None and not summary_all.empty:
            for prod, prod_df in summary_all.groupby("Product"):
                prod_df_sorted = prod_df.sort_values("Question")
                _write_product_pie_charts(wb, str(prod), prod_df_sorted)

    print(f"[ok] Wrote Excel report to {out_path}")


# ---------------- Main ----------------

def main():
    load_dotenv()

    parser = argparse.ArgumentParser(
        description="Analyze customer survey answers and export a wide-format Excel report with per-question pie charts."
    )
    parser.add_argument("--input", required=True, help="Path to input CSV.")
    parser.add_argument("--industry", required=True, help="Industry context, e.g., 'Fashion'.")
    parser.add_argument("--output", required=False, help="Output Excel path. Defaults to 'data analysis output.xlsx'")
    args = parser.parse_args()

    # Client selection (Demo vs OpenAI)
    api_key = os.getenv("OPENAI_API_KEY")
    client = None
    if api_key:
        if OpenAI is None:
            print("[error] openai package not installed. See requirements.txt", file=sys.stderr)
            sys.exit(1)
        client = OpenAI(api_key=api_key)
    else:
        print("[info] OPENAI_API_KEY not set. Running in Demo Mode with offline analyzer.", file=sys.stderr)

    # Read CSV
    try:
        df = pd.read_csv(args.input)
    except FileNotFoundError:
        print(f"[error] File not found: {args.input}", file=sys.stderr); sys.exit(1)
    except Exception as e:
        print(f"[error] Could not read CSV: {e}", file=sys.stderr); sys.exit(1)

    if df.shape[1] < 4:
        print("[error] Need at least 4 columns: Email, Name, Products, and one question column.", file=sys.stderr)
        sys.exit(1)

    out_path = args.output or "data analysis output.xlsx"

    wide = analyze_dataframe_wide(df, args.industry, client)
    write_excel_wide(wide, out_path)


if __name__ == "__main__":
    main()